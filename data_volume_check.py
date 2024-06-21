import os

import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# # 定义除法运算函数
# def safe_divide(numerator, denominator):
#     return np.where(denominator == 0, '#DIV/0!', numerator / denominator)


# 读取Excel文件中的数据
file1 = 'data_volume_check/File1.xlsx'
file2 = 'data_volume_check/File2.xlsx'

# 假设Sheet1是工作表的名称
df1 = pd.read_excel(file1, sheet_name='C360-2.0')
df2 = pd.read_excel(file2, sheet_name='c360')

# 创建辅助列（拼接A列和B列）  pandas 默认读取第一行作为列名
df1['Key'] = df1['data_source_name'].astype(str) + df1['table_name'].astype(str)
df2['Key'] = df2['data_source_name'].astype(str) + df2['table_name'].astype(str)

# 选择需要合并的列，包括辅助列
df2_subset = df2[['Key', 'source_data_volumn', 'target_data_volumn']]

# 合并数据
merged_df = pd.merge(df1, df2_subset, on='Key', how='left')

# 删除辅助列
merged_df.drop(columns=['Key'], inplace=True)

# 将NaN替换为#N/A
merged_df = merged_df.fillna('#N/A')

# 定义一个函数来处理指定的列
# def process_specified_columns(df):
#     # 添加新列E和F，并初始化
#     df['Compare'] = 0
#     df['Increase%'] = 0
#
#     # 处理Compare
#     def process_Compare(row):
#         cur_col = -2  # 当前列
#         # if(safe_divide(row[cur_col-1]-row[cur_col-2]  , row[cur_col-1]-row[cur_col-2]) == 0):
#
#         return 1 if safe_divide(row[cur_col - 1] - row[cur_col - 3], row[cur_col - 3] - row[cur_col - 5]) > 1.5 else 0
#
#     # 处理Increase列
#     def process_Increase(row):
#         cur_col = -1  # 当前列
#         return (row[cur_col-2] - row[cur_col-4]) / row[cur_col-2]
#
#     # 应用函数到E列
#     df['Compare'] = df.apply(lambda row: process_Compare(row), axis=1)
#
#     # 应用函数到F列
#     df['Increase%'] = df.apply(lambda row: process_Increase(row), axis=1)
#
#     return df


# # 应用处理函数
# merged_df = process_specified_columns(merged_df)

# s输出前先对 列名更改一下

output_file = 'data_volume_check/Merged_File.xlsx'

# 如果文件存在，则删除它
if os.path.exists(output_file):
    os.remove(output_file)

# 保存结果到新的Excel文件
merged_df.to_excel(output_file, index=False)

# wb = load_workbook(output_file)
# ws = wb.active # 获取当前活动的工作表
#
# # 获取列数  也就是当前
# total_columns = ws.max_column

# 读取output_file
output_df = pd.read_excel('data_volume_check/Merged_File.xlsx')
# 在顶部插入一行，将所有其他行下移
output_df = pd.concat([pd.DataFrame([output_df.columns.tolist()], columns=output_df.columns), output_df],
                      ignore_index=True)
# 获取output_file的最后两列
last_two_columns = output_df.iloc[:, -2:]

# 读取file1的C360-2.0工作表
file1_df = pd.read_excel('data_volume_check/File1.xlsx', sheet_name='C360-2.0')

# 将最后两列添加到C360-2.0工作表的末尾
file1_df = pd.concat([file1_df, last_two_columns], axis=1)

# 获取新添加的列的列名
new_column_names = file1_df.columns[-2:]

# 将新添加的列的列名设置为空
file1_df.rename(columns={name: '' for name in new_column_names}, inplace=True)
# 保存结果到新的Excel文件
file1_df.to_excel('data_volume_check/File1.xlsx', sheet_name='C360-2.0', index=False)

## 有 bug  暂时自己在 excel  里  编写函数拉取吧
# wb = load_workbook('data_volume_check/File1.xlsx')
# ws = wb['C360-2.0']  # 获取C360-2.0工作表
# # 获取当前最后一列的索引
# last_column_index = ws.max_column
# print(last_column_index)
#
# last_column_letter = get_column_letter(last_column_index)
#
# # for row in range(2, ws.max_row + 1):
# #     ws[f'{last_column_index+1}{row}'] = f'=IF((last_column_index{row} - {last_column_index-2}{row})/({last_column_index-2}{row} - {last_column_index-4}{row}))>1.5,1,0)'
# #
# #     ws[f'{last_column_index+2}{row}'] = f'=(last_column_index{row} - {last_column_index-2}{row})/last_column_index{row}'
#
# for row in range(2, ws.max_row + 1):
#     ws[f'{get_column_letter(last_column_index+1)}{row}'] = f'=IF(({get_column_letter(last_column_index)}{row} - {get_column_letter(last_column_index-2)}{row})/({get_column_letter(last_column_index-2)}{row} - {get_column_letter(last_column_index-4)}{row}))>1.5,1,0)'
#     ws[f'{get_column_letter(last_column_index+2)}{row}'] = f'=({get_column_letter(last_column_index)}{row} - {get_column_letter(last_column_index-2)}{row})/{get_column_letter(last_column_index)}{row}'
#
#
# wb.save('data_volume_check/test.xlsx')
