import pandas as pd
from openpyxl import load_workbook


# 读取初始表和参数表
tb1 = pd.read_excel('source/tb1.xlsx')
source1 = pd.read_excel('source/source1.xlsx')
source2 = pd.read_excel('source/source2.xlsx')
source3 = pd.read_excel('source/source3.xlsx')
source4 = pd.read_excel('source/source4.xlsx')

cols_tol = tb1.shape[1] # 列数
# tb1 添加 10 列（列名从'col1'到'col10'）
for i in range(1, 11):
    tb1[f'col{i}'] = None

# 提取四张参数表中的 '521id' 字段并分别命名
tb1.rename(columns={'col7': 'rep'}, inplace=True)
tb1.rename(columns={'col8': 'flm'}, inplace=True)
tb1.rename(columns={'col9': 'bm'}, inplace=True)
tb1.rename(columns={'col10': 'ha'}, inplace=True)
tb1.rename(columns={'col2': 'rep requested'}, inplace=True)
tb1.rename(columns={'col3': 'flm requested'}, inplace=True)
tb1.rename(columns={'col4': 'bm requested'}, inplace=True)
tb1.rename(columns={'col5': 'ha requested'}, inplace=True)

tb1.rename(columns={'col1': ''}, inplace=True)  # 重命名列名
tb1.rename(columns={'col6': ''}, inplace=True)  # 重命名列名

# 获取源表的最大行索引
max_index = max([source1.index.max(), source2.index.max(), source3.index.max(), source4.index.max()])

# 扩展主表的行数
tb1 = tb1.reindex(range(max_index + 1))


# 将四张参数表中的 '521id' 字段添加到 tb1 表中
# 实际表中好像不叫这个到时候改一下
tb1['rep'] = source1['521id']
tb1['flm'] = source2['521id']
tb1['bm'] = source3['521id']
tb1['ha'] = source4['521id']

# 将提取的 '521id' 字段移到表的最后四列
# columns_order = tb1.columns.tolist()
# columns_order = columns_order[:len(columns_order) - 4] + ['rep', 'flm', 'bm', 'ha']
# tb1 = tb1[columns_order]

# 添加空的六列（列名从'col11'到'col16'）
# for i in range(11, 17):
#     tb1[f'col{i}'] = None

# 保存临时文件以便于使用openpyxl
temp_filename = 'out/temp_output.xlsx'
tb1.to_excel(temp_filename, index=False)

# 使用openpyxl插入VLOOKUP公式
wb = load_workbook(temp_filename)  # 打开临时文件
ws = wb.active

# 获取列字母
def get_column_letter(col_num):
    letter = ''
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

# 获取各列的列字母
col_rep = get_column_letter(tb1.columns.get_loc('rep') + 1)
col_flm = get_column_letter(tb1.columns.get_loc('flm') + 1)
col_bm = get_column_letter(tb1.columns.get_loc('bm') + 1)
col_ha = get_column_letter(tb1.columns.get_loc('ha') + 1)
col_521id = get_column_letter(tb1.columns.get_loc('521id') + 1)
col2 = get_column_letter(tb1.columns.get_loc('rep requested') + 1)
col3 = get_column_letter(tb1.columns.get_loc('flm requested') + 1)
col4 = get_column_letter(tb1.columns.get_loc('bm requested') + 1)
col5 = get_column_letter(tb1.columns.get_loc('ha requested') + 1)

# 这里要改  改成直接定位到列

# 插入VLOOKUP公式到第11到14列
for row in range(2, ws.max_row + 1):
    ws[f'{col2}{row}'] = f'=VLOOKUP(C{row}, ${col_rep}$2:${col_rep}${ws.max_row}, 1, FALSE)'
    ws[f'{col3}{row}'] = f'=VLOOKUP(C{row}, ${col_flm}$2:${col_flm}${ws.max_row}, 1, FALSE)'
    ws[f'{col4}{row}'] = f'=VLOOKUP(C{row}, ${col_bm}$2:${col_bm}${ws.max_row}, 1, FALSE)'
    ws[f'{col5}{row}'] = f'=VLOOKUP(C{row}, ${col_ha}$2:${col_ha}${ws.max_row}, 1, FALSE)'


# 添加sheet分页
# ws = wb.create_sheet('验证')

# 保存最终文件
final_filename = 'out/output.xlsx'
wb.save(final_filename)

print("成功输出为output.xlsx")

